#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import xlsxwriter
import xlrd

# ----------读取xlsx-----------
excelPath = "F:/code/python/test.xlsx"
# 加载xlsx
workBook = load_workbook(excelPath)
# 仅仅读取Sheet1
workSheet = workBook.get_sheet_by_name("Sheet1")

# 行、列的初始值
rownum = 1
columnnum = 1

# 获取第一行第一列单元格的值
cell = workSheet.cell(row=rownum, column=columnnum).value
print(cell)

# ----------写入xlsx-----------
# 打开某个excel
with xlsxwriter.Workbook("F:/code/python/test1.xlsx") as workbook:
    # 设置Sheet的名字为haha
    worksheet = workbook.add_worksheet("haha")
    # 行、列的初始值
    rownum = 0
    columnnum = 0
    # 依次写入每个单元格
    worksheet.write(rownum, columnnum, "one")
workbook.close()

# ----------读取xls-----------
excelPath = "F:/code/python/test.xls"
# 加载xlsx
wb = xlrd.open_workbook(excelPath)
# 仅仅读取Sheet1
ws = wb.sheet_by_name("Sheet1")
# 行、列的初始值
rownum = 0
columnnum = 0
# 获取第一行第一列单元格的值
cel = ws.cell(rowx=rownum, colx=columnnum).value
print(cel)
